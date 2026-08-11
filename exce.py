import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.common.by import By

# Creating Excel
def create_test_excel():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Logins"

    sheet.cell(row = 1, column = 1, value = "Username")
    sheet.cell(row = 1, column = 2, value = "Password")

    sheet.cell(row = 2, column = 1, value = "standard_user")
    sheet.cell(row = 2, column = 2, value = "secret_sauce")

    sheet.cell(row = 3, column = 1, value = "locked_out_user")
    sheet.cell(row = 3, column = 2, value = "secret_sauce")

    workbook.save("test_data.xlsx")

create_test_excel()

# Reading Excel
workbook = openpyxl.load_workbook("test_data.xlsx")
sheet = workbook["Logins"]

service = Service(GeckoDriverManager().install())
driver = webdriver.Firefox(service = service)

try:
    for row in range(2, sheet.max_row + 1):
        
        excel_username = sheet.cell(row = row, column = 1).value
        excel_password = sheet.cell(row = row, column = 2).value

        driver.get("https://www.saucedemo.com")
        driver.find_element(By.ID, "user-name").send_keys(excel_username)
        driver.find_element(By.ID, "password").send_keys(excel_password)
        time.sleep(5)
        driver.find_element(By.ID, "login-button").click()
        time.sleep(2)

finally: 
    driver.quit()